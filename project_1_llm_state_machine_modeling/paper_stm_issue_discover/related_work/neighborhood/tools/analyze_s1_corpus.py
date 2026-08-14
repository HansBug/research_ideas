"""在第三方 LLM4MDE 映射研究（S1）的 86 篇抽取表上做交叉统计。

⭐ **为什么这份数据值钱**：S1（*Large Language Models in Model-Driven Engineering:
A Systematic Mapping Study*，EMSE，DOI `10.1007/s10664-026-10921-4`）把 86 篇 primary
study 逐篇抽成了 **43 个字段**并公开了复现包。⭐ 那是一份**第三方标注的、带分母的**
形态数据 —— ⛔ 比我们自己临时数几十篇可靠得多。

⭐ **它回答我们自己数不出来的问题**：`Model Validation` 这一类里到底有几篇的制品是
**行为模型**（S1 自己没有交叉列出这一格）。

## 取数据

⛔ **`g5by9` 是匿名 view-only 项目，鉴权靠 query string。** ⚠️ 掐掉 `?view_only=`
去请求必然 401 —— ⛔ 而 401 与「需要登录」在状态码上**完全无法区分**，⭐ 主 session
第一次就是这么误判成「取不到」的。token 逐字写在 S1 PDF 的参考文献第 112 条里::

    T=5c10c1e56be3480d8d25e017b4276f7a
    curl -sL "https://osf.io/download/8vpkj/?view_only=$T" -o S1_data_extraction.xlsx
    # sha256 5a396fe4e3c1b5e292342469172106349dbfb464da783d0a4b91cb31b1e67279  (76325 B)

⚠️ sheet `framework`，**表头在第 2 行**（⛔ 不是第 1 行）；⭐ 989 行里 **86 行有
`Publication_Year`**，那 86 行才是 primary study，其余是空壳并入行。

用法::

    python -m tools.analyze_s1_corpus --xlsx /tmp/l3/assets/S1_data_extraction.xlsx
    python -m tools.analyze_s1_corpus --xlsx ... --list-behavioral
"""

from __future__ import annotations

import argparse
import collections
import re
import sys
from pathlib import Path

#: ⭐ 用来判「这篇的制品是不是行为模型」的词。
#:
#: ⛔⛔ **`automat` 这个前缀被刻意排除，⭐ 尽管它看起来是 `automaton` 的词根。**
#: ⚠️ 实测：全表 `automat\w*` 共 11 次命中，⭐ 其中只有 **1 次**是 `automata`，
#: ⛔ 其余 10 次是 `Automation` / `automating` / `automatic` / `automated` /
#: `AutomationML` —— ⭐ 即**十分之九是「自动化」不是「自动机」**。
#:
#: ⭐ 这是 `CLAUDE.md` §11 点名的那类错误：⛔ **用词法判据冒充语义判断**。
#: ⭐ 后果是 `P059`（*Evaluating the Quality of Class **Diagrams***，⛔ 类图，
#: 与行为模型无关）因为标题里有 `Automation` 而被判成行为类。
#:
#: ⭐ 故这里只收 `automaton` / `automata` 的**完整词形**。
_BEHAVIORAL = re.compile(
    r"state\s*machine|statechart|state\s+chart|state\s+diagram"
    r"|\bFSM\b|\bEFSM\b|\bHSM\b"
    r"|activity\s+diagram|sequence\s+diagram|interaction\s+diagram"
    r"|\bBPMN\b|workflow|process\s+model"
    r"|petri|event-b|\bCSP\b|\bLTS\b|\bSCXML\b|Stateflow"
    r"|\bautomaton\b|\bautomata\b"          #: ⛔ 完整词，不用 `automat` 前缀
    r"|behaviou?ral\s+model|scenario-based",
    re.I,
)

#: ⛔ 这些字段用来判制品类型。⚠️ **`Titles` 不在其中** —— ⭐ 标题里出现某个词
#: 不代表那是本文的制品（`P050` 的输入里提到 activity diagram，⛔ 但它做的是
#: 转换工具推荐，输出是 relevance scores）。
_ARTIFACT_FIELDS = (
    "Modeling_Language_Details",
    "Output_Artifact_Type",
    "Output_Artifact_Categories",
)


def load(xlsx: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError:  # pragma: no cover
        print("⛔ 需要 openpyxl：pip install openpyxl", file=sys.stderr)
        raise
    ws = openpyxl.load_workbook(xlsx, data_only=True)["framework"]
    hdr = [c.value for c in ws[2]]  #: ⚠️ 表头在第 2 行
    rows = [dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=3)]
    #: ⭐ 只留 primary study。⛔ 不能用「有 ID」筛 —— 空壳并入行也有 ID。
    return [r for r in rows if r.get("Publication_Year")]


def artifact_blob(row: dict) -> str:
    return " | ".join(str(row.get(k) or "") for k in _ARTIFACT_FIELDS)


def is_behavioral(row: dict) -> tuple[bool, list[str]]:
    """⭐ 同时返回命中词，⛔ 让人能复核为什么判成行为类。"""
    hits = _BEHAVIORAL.findall(artifact_blob(row))
    return bool(hits), sorted({h.strip().lower() for h in hits})


def multi(rows: list[dict], col: str) -> collections.Counter:
    """⚠️ 多值列用 `,` 与 `;` 混合分隔，⛔ 只按其一切会少数。"""
    c: collections.Counter = collections.Counter()
    for r in rows:
        raw = str(r.get(col) or "")
        for part in (p.strip() for p in raw.replace(";", ",").split(",")):
            if part:
                c[part] += 1
    return c


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", type=Path, required=True)
    ap.add_argument("--list-behavioral", action="store_true")
    args = ap.parse_args(argv)

    rows = load(args.xlsx)
    print(f"# S1 语料交叉统计（primary study **{len(rows)}** 篇）\n")
    if len(rows) != 86:
        print(f"⚠️ ⛔ 期望 86 篇，实得 {len(rows)} —— 表结构可能变了，下面的数字不要直接引用。\n")

    print("## 1. MDE 任务分布（⭐ 与 PDF Fig. 2 对拍）\n")
    print("| 任务 | 篇数 |\n| :-- | --: |")
    for k, v in multi(rows, "MDE_Task_Type").most_common():
        print(f"| {k} | {v} |")

    beh = [r for r in rows if is_behavioral(r)[0]]
    print(f"\n## 2. ⭐ 行为类制品子集：**{len(beh)} / {len(rows)}**\n")
    print("⛔ 判据只看制品字段（`Modeling_Language_Details` / `Output_Artifact_Type` /")
    print("`Output_Artifact_Categories`），⛔ **不看标题、不看输入** —— ⚠️ 输入里提到某种图")
    print("不代表本文的产物是那种图。\n")

    print("### 2.1 ⭐⭐ 任务 × 行为类交叉（⛔ S1 自己没列这一格）\n")
    print("| 任务 | 全部 | ⭐ 其中行为类 |\n| :-- | --: | --: |")
    for task, n in multi(rows, "MDE_Task_Type").most_common():
        nb = sum(1 for r in beh if task in (r.get("MDE_Task_Type") or ""))
        star = " ⭐⭐" if task == "Model Validation" else ""
        print(f"| {task}{star} | {n} | **{nb}** |")

    mvb = [r for r in beh if "Model Validation" in (r.get("MDE_Task_Type") or "")]
    print(f"\n⭐⭐ **`Model Validation` ∩ 行为类 = {len(mvb)} / {len(rows)}**\n")
    for r in mvb:
        _, hits = is_behavioral(r)
        print(f"- `{r.get('ID')}` **{str(r.get('Titles'))[:70]}** · {r.get('Publication_Year')} · 命中 {hits}")

    print("\n## 3. ⭐ 行为类子集 vs 全语料（⛔ 看它是不是更成熟）\n")
    print("| 维度 | 全语料 | 行为类子集 |\n| :-- | :-- | :-- |")
    for col in ("Autonomy_Level", "Baseline_Comparison", "Human_Evaluation_Included", "Cost_Efficiency_Evaluated"):
        a = multi(rows, col).most_common(3)
        b = multi(beh, col).most_common(3)
        fa = " · ".join(f"{k} {v}" for k, v in a) or "—"
        fb = " · ".join(f"{k} {v}" for k, v in b) or "—"
        print(f"| `{col}` | {fa} | {fb} |")

    print("\n## 4. ⭐ 执行结构（行为类子集）\n")
    print("| 结构 | 篇数 |\n| :-- | --: |")
    for k, v in multi(beh, "Execution_Structure").most_common():
        print(f"| {k} | {v} |")

    if args.list_behavioral:
        print(f"\n## 5. 行为类 {len(beh)} 篇明细\n")
        print("| ID | 年 | 任务 | 制品 | 命中词 |\n| :-- | :-: | :-- | :-- | :-- |")
        for r in sorted(beh, key=lambda r: str(r.get("ID"))):
            _, hits = is_behavioral(r)
            print(
                f"| `{r.get('ID')}` | {str(r.get('Publication_Year')).replace('.0','')} "
                f"| {str(r.get('MDE_Task_Type'))[:28]} | {str(r.get('Modeling_Language_Details'))[:34]} | {','.join(hits)} |"
            )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
