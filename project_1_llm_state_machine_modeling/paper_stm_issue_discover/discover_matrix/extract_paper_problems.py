"""Extract the source paper's own per-case problem record from its public workbook.

Why this exists
---------------
The 47 expected issues in issue #166 are an LLM agent's open-ended induction from
`<NL, author PlantUML, FCSTM>`.  The *paper* has its own record of what is wrong
with each of the 60 generated models -- three columns of hallucinations per case,
each with a `Resolved` flag, plus Phase-I/Phase-II F1.  That record is the only
external anchor available for asking "what did our ground truth miss", so it has
to be extractable on demand rather than transcribed once by hand.

Source: `Experiment Results.xlsx`, sheet `STM Results` (63 rows x 38 columns), from
the paper's public Drive folder, vendored under
`paper_stm_issue_discover/corpora/seed_library/llms-emp-stm-subset/assets/raw/drive_download/`.

Column map (1-indexed), read off the two header rows:

    A  Model Source          H  LLMs                    I  Generation PlantUML
    T  Format Hallucinations V  Resolved                 U  Result with Format Checking
    Y  SysML Grammar Halluc. AA Resolved                 Z  Result with Grammar Checking
    AD Semmantic Halluc.     AG Resolved                 AE Result with Semantic Checking
    S  F1 (Phase-I)          AK F1 (Phase-II)

`Model Source` and `Model Name` are merged cells spanning each model's six LLM
rows, so they are carried forward.  Rows are keyed to our case ids by
`(LLM, normalised domain)` against the frozen ledger -- the workbook orders rows
by (LLM, model) and our ids follow the same order, but matching on content rather
than position is what makes the 60/60 alignment checkable.  Whitespace has to be
normalised: the microwave domain carries an embedded newline in one of the two
sources and would otherwise fail to match on six rows.

Usage:
    PYTHONPATH=<repo root> python extract_paper_problems.py [out.json]
"""

from __future__ import annotations

import json
import pathlib
import re
import sys

HERE = pathlib.Path(__file__).resolve().parent
ROOT = HERE.resolve().parents[2]
WORKBOOK = (
    ROOT
    / "project_1_llm_state_machine_modeling/paper_stm_issue_discover/corpora/seed_library"
    / "llms-emp-stm-subset/assets/raw/drive_download/Experiment Results.xlsx"
)
LEDGER = (
    ROOT / ".omx/specs/autoresearch-paper1-llms-emp-60-expected-issues/ledger.json"
)

COL = {
    "src": 1, "name": 2, "llm": 8, "gen": 9,
    "pl_acc": 12, "gr_acc": 14, "f1_phase1": 19,
    "fmt": 20, "fmt_resolved": 22,
    "gram": 25, "gram_resolved": 27,
    "sem": 30, "sem_resolved": 33,
    "tp2": 34, "fp2": 35, "fn2": 36, "f1_phase2": 37,
}

_BLANK = {None, "None", ""}


def _norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value)).strip()


def _text(value: object) -> str | None:
    return _norm(value) if value not in _BLANK else None


def extract() -> dict[str, dict]:
    import openpyxl

    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    sheet = workbook["STM Results"]
    rows: list[dict] = []
    source = name = None
    for index in range(2, sheet.max_row + 1):
        cell = lambda key: sheet.cell(index, COL[key]).value
        # Merged cells report their value only on the first row of the span.
        source = _text(cell("src")) or source
        name = _text(cell("name")) or name
        if not cell("llm") or not cell("gen"):
            continue
        rows.append({
            "workbook_row": index,
            "model_source": source,
            "model_name": name,
            "llm": _norm(cell("llm")),
            "plantuml_accuracy": cell("pl_acc"),
            "grammar_accuracy": cell("gr_acc"),
            "f1_phase1": cell("f1_phase1"),
            "f1_phase2": cell("f1_phase2"),
            "format_hallucinations": _text(cell("fmt")),
            "format_resolved": cell("fmt_resolved"),
            "grammar_hallucinations": _text(cell("gram")),
            "grammar_resolved": cell("gram_resolved"),
            "semantic_hallucinations": _text(cell("sem")),
            "semantic_resolved": cell("sem_resolved"),
        })
    workbook.close()

    ledger = json.loads(LEDGER.read_text())
    by_key = {
        (case["llm"], _norm(case["domain"])): case["case_id"]
        for case in ledger["cases"]
    }
    mapped: dict[str, dict] = {}
    unmapped: list[tuple[str, str]] = []
    for row in rows:
        case_id = by_key.get((row["llm"], _norm(row["model_name"])))
        if case_id is None:
            unmapped.append((row["llm"], row["model_name"]))
        else:
            mapped[case_id] = row
    if unmapped or len(mapped) != 60:
        raise SystemExit(
            f"alignment failed: {len(mapped)}/60 mapped, unmapped={unmapped}. "
            "Refusing to emit a partial record -- a missing case reads as a case "
            "with no reported problems."
        )
    return mapped


def main() -> int:
    out = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 else HERE / "paper_reported_problems.json"
    mapped = extract()
    out.write_text(json.dumps(mapped, ensure_ascii=False, indent=1) + "\n")
    kinds = {
        "format": sum(1 for r in mapped.values() if r["format_hallucinations"]),
        "grammar": sum(1 for r in mapped.values() if r["grammar_hallucinations"]),
        "semantic": sum(1 for r in mapped.values() if r["semantic_hallucinations"]),
    }
    print(f"wrote {len(mapped)} cases -> {out}")
    print(f"  cases with reported problems: {kinds}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
