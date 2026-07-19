#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _repo_root() -> Path:
    for parent in Path(__file__).resolve().parents:
        if (parent / "project_1_llm_state_machine_modeling").is_dir():
            return parent
    raise RuntimeError("repository root not found")


REPO_ROOT = _repo_root()
PAPER_ROOT = REPO_ROOT / "project_1_llm_state_machine_modeling/paper_stm_repair"
DEFAULT_WORKBOOK = (
    PAPER_ROOT
    / "corpora/seed_library/llms-emp-stm-subset/assets/raw/drive_download"
    / "Experiment Results.xlsx"
)
DEFAULT_OUTPUT = (
    PAPER_ROOT
    / "corpora/seed_library/llms-emp-stm-subset/assets/extracted"
    / "feedback_final_pairs.jsonl"
)
DEFAULT_SUMMARY = DEFAULT_OUTPUT.with_name("feedback_final_validation_summary.json")

SHEET = "STM Results"
EXPECTED_PAIR_COUNT = 60
PAIR_SCHEMA_VERSION = "llms_emp.feedback_final_pair.v2"
SUMMARY_SCHEMA_VERSION = "llms_emp.feedback_final_pool.v2"
SELECTION_POLICY_ID = "author_phase_ii_last_nonblank.v1"

STAGES = (
    {
        "stage_id": "phase_i_generation",
        "output": "Generation PlantUML",
        "feedback": None,
        "resolved": None,
        "resolved_offset": None,
    },
    {
        "stage_id": "phase_ii_format",
        "output": "Result with Format Checking",
        "feedback": "Format Hallucinations",
        "resolved": "Resolved",
        "resolved_offset": 1,
    },
    {
        "stage_id": "phase_ii_grammar",
        "output": "Result with Grammar Checking",
        "feedback": "SysML Grammar Hallucinations",
        "resolved": "Resolved",
        "resolved_offset": 1,
    },
    {
        "stage_id": "phase_ii_semantic",
        "output": "Result with Semantic Checking",
        "feedback": "Semmantic Hallucinations",
        "resolved": "Resolved",
        "resolved_offset": 2,
    },
)
FINAL_STAGE_COLUMNS = tuple(stage["output"] for stage in reversed(STAGES))
LINEAGE_COLUMNS = tuple(stage["output"] for stage in STAGES)
IDENTITY_COLUMNS = (
    "Model Source",
    "Model Name",
    "Requirement Description",
    "PlantUML",
    "LLMs",
)


def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _sha256_text(text: str) -> str:
    return _sha256_bytes(text.encode("utf-8"))


def _canonical_json(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def _repo_relative(path: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(REPO_ROOT))
    except ValueError:
        return str(resolved)


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _required_headers() -> tuple[str, ...]:
    headers = list(IDENTITY_COLUMNS)
    for stage in STAGES:
        headers.append(stage["output"])
        if stage["feedback"] is not None:
            headers.append(stage["feedback"])
    return tuple(dict.fromkeys(headers))


def _header_schema(headers: list[object]) -> list[dict[str, Any]]:
    return [
        {
            "column_index": index,
            "column_letter": get_column_letter(index),
            "header": header,
        }
        for index, header in enumerate(headers, start=1)
    ]


def _header_positions(headers: list[object]) -> dict[str, list[int]]:
    positions: dict[str, list[int]] = defaultdict(list)
    for index, header in enumerate(headers):
        if header is not None:
            positions[str(header)].append(index)
    missing = [header for header in _required_headers() if not positions[header]]
    if missing:
        raise RuntimeError(f"workbook missing required columns: {missing}")
    ambiguous = {
        header: [get_column_letter(index + 1) for index in positions[header]]
        for header in _required_headers()
        if len(positions[header]) != 1
    }
    # Resolved is intentionally repeated once per Phase-II stage and is bound
    # positionally below rather than through this header map.
    ambiguous.pop("Resolved", None)
    if ambiguous:
        raise RuntimeError(f"workbook has ambiguous required columns: {ambiguous}")
    return positions


def _column_index(positions: dict[str, list[int]], header: str) -> int:
    return positions[header][0]


def _cell_descriptor(
    *,
    header: str | None,
    index: int | None,
    excel_row: int,
    value: object,
    include_value: bool,
) -> dict[str, Any]:
    if header is None or index is None:
        return {
            "header": None,
            "column_letter": None,
            "cell": None,
            "present": False,
            "value_type": "not_applicable",
            "sha256": None,
            "char_length": 0,
            "byte_length": 0,
            "value": None,
        }
    present = not _is_blank(value)
    text = value if isinstance(value, str) else None
    descriptor = {
        "header": header,
        "column_letter": get_column_letter(index + 1),
        "cell": f"{get_column_letter(index + 1)}{excel_row}",
        "present": present,
        "value_type": "null" if value is None else type(value).__name__,
        "sha256": _sha256_text(text) if present and text is not None else None,
        "char_length": len(text) if present and text is not None else 0,
        "byte_length": len(text.encode("utf-8")) if present and text is not None else 0,
    }
    if include_value:
        descriptor["value"] = value
    return descriptor


def _resolved_index(stage_index: int, output_index: int, headers: list[object]) -> int:
    if stage_index == 0:
        raise ValueError("Phase-I has no Resolved cell")
    expected = output_index + int(STAGES[stage_index]["resolved_offset"])
    if expected >= len(headers) or headers[expected] != "Resolved":
        raise RuntimeError(
            f"{STAGES[stage_index]['output']} must be followed by Resolved"
        )
    return expected


def _has_plantuml_envelope(text: str) -> bool:
    # The author workbook contains one output with a trailing quote after
    # @enduml. It is still an identifiable PlantUML payload and is preserved
    # byte-for-byte; this check rejects non-code sentinels, not syntax defects.
    return bool(
        re.match(r"^\s*@startuml\b", text, flags=re.IGNORECASE)
        and re.search(r"(?m)^\s*@enduml\b[\"']?\s*$", text, flags=re.IGNORECASE)
    )


def _stage_lineage(
    source_row: tuple[object, ...],
    *,
    headers: list[object],
    positions: dict[str, list[int]],
    excel_row: int,
) -> list[dict[str, Any]]:
    lineage: list[dict[str, Any]] = []
    for stage_index, stage in enumerate(STAGES):
        output_index = _column_index(positions, stage["output"])
        output_value = source_row[output_index]
        output = _cell_descriptor(
            header=stage["output"],
            index=output_index,
            excel_row=excel_row,
            value=output_value,
            include_value=False,
        )
        if output["present"]:
            if not isinstance(output_value, str) or not _has_plantuml_envelope(
                output_value
            ):
                raise RuntimeError(
                    f"STM Results {output['cell']} has a non-PlantUML stage output"
                )

        feedback_header = stage["feedback"]
        feedback_index = (
            _column_index(positions, feedback_header)
            if feedback_header is not None
            else None
        )
        feedback_value = (
            source_row[feedback_index] if feedback_index is not None else None
        )
        resolved_index = (
            _resolved_index(stage_index, output_index, headers)
            if stage_index > 0
            else None
        )
        resolved_value = (
            source_row[resolved_index] if resolved_index is not None else None
        )
        lineage.append(
            {
                "stage_id": stage["stage_id"],
                "output": output,
                "feedback": _cell_descriptor(
                    header=feedback_header,
                    index=feedback_index,
                    excel_row=excel_row,
                    value=feedback_value,
                    include_value=True,
                ),
                "resolved": _cell_descriptor(
                    header="Resolved" if resolved_index is not None else None,
                    index=resolved_index,
                    excel_row=excel_row,
                    value=resolved_value,
                    include_value=True,
                ),
            }
        )
    return lineage


def _read_workbook_rows(
    workbook_path: Path,
) -> tuple[list[object], dict[str, list[int]], list[tuple[int, tuple[object, ...]]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        if SHEET not in workbook.sheetnames:
            raise RuntimeError(f"missing workbook sheet: {SHEET}")
        source_rows = workbook[SHEET].iter_rows(values_only=True)
        headers = list(next(source_rows))
        positions = _header_positions(headers)
        nl_index = _column_index(positions, "Requirement Description")
        rows = [
            (excel_row, tuple(row))
            for excel_row, row in enumerate(source_rows, start=2)
            if not _is_blank(row[nl_index])
        ]
        return headers, positions, rows
    finally:
        workbook.close()


def verify_feedback_final_trace(
    workbook_path: Path,
    extracted: list[dict[str, Any]],
) -> dict[str, Any]:
    """Independently re-read the workbook and verify every extracted locator."""

    workbook_path = workbook_path.resolve()
    workbook_sha256 = _sha256_bytes(workbook_path.read_bytes())
    headers, positions, source_rows = _read_workbook_rows(workbook_path)
    schema = _header_schema(headers)
    schema_sha256 = _sha256_text(_canonical_json(schema))
    if len(extracted) != len(source_rows):
        raise RuntimeError(
            f"feedback-final trace row count mismatch: {len(extracted)} != {len(source_rows)}"
        )

    for data_row_index, ((excel_row, source_row), pair) in enumerate(
        zip(source_rows, extracted)
    ):
        pair_id = f"llms_emp_feedback_final_{data_row_index:04d}"
        if pair.get("schema_version") != PAIR_SCHEMA_VERSION:
            raise RuntimeError(f"{pair_id} pair schema_version mismatch")
        if pair.get("selection_policy_id") != SELECTION_POLICY_ID:
            raise RuntimeError(f"{pair_id} selection policy mismatch")
        if pair.get("pair_id") != pair_id:
            raise RuntimeError(f"feedback-final pair_id mismatch at row {data_row_index}")
        if pair.get("source_sha256") != workbook_sha256:
            raise RuntimeError(f"{pair_id} workbook hash mismatch")
        if pair.get("workbook_schema_sha256") != schema_sha256:
            raise RuntimeError(f"{pair_id} workbook schema fingerprint mismatch")
        if pair.get("source_row_index") != data_row_index:
            raise RuntimeError(f"{pair_id} source_row_index mismatch")
        if pair.get("source_excel_row") != excel_row:
            raise RuntimeError(f"{pair_id} source_excel_row mismatch")

        lineage = _stage_lineage(
            source_row,
            headers=headers,
            positions=positions,
            excel_row=excel_row,
        )
        if pair.get("stage_lineage") != lineage:
            raise RuntimeError(f"{pair_id} stage lineage mismatch")
        selected = next(
            stage for stage in reversed(lineage) if stage["output"]["present"]
        )
        selected_index = _column_index(positions, selected["output"]["header"])
        selected_text = source_row[selected_index]
        generation_index = _column_index(positions, "Generation PlantUML")
        generation_text = source_row[generation_index]
        identity = {
            "model_source": ("Model Source", "model_source"),
            "model_name": ("Model Name", "model_name"),
            "llm": ("LLMs", "llm"),
            "nl_text": ("Requirement Description", "nl_text"),
        }
        for _, (header, key) in identity.items():
            if pair.get(key) != source_row[_column_index(positions, header)]:
                raise RuntimeError(f"{pair_id} {key} workbook mismatch")
        expected = {
            "selected_stage": selected["stage_id"],
            "selected_stage_column": selected["output"]["header"],
            "selected_stage_cell": selected["output"]["cell"],
            "stm0_source_column": selected["output"]["header"],
            "stm0_text": selected_text,
            "stm0_sha256": _sha256_text(selected_text),
            "phase_i_stm0_sha256": _sha256_text(generation_text),
            "phase_i_changed": selected_text != generation_text,
            "is_postprocessed": selected["stage_id"] != "phase_i_generation",
            "is_phase_i_fallback": selected["stage_id"] == "phase_i_generation",
        }
        for key, value in expected.items():
            if pair.get(key) != value:
                raise RuntimeError(f"{pair_id} {key} trace mismatch")

        source_cells = {
            key: f"{get_column_letter(_column_index(positions, header) + 1)}{excel_row}"
            for key, header in {
                "model_source": "Model Source",
                "model_name": "Model Name",
                "nl": "Requirement Description",
                "reference_plantuml": "PlantUML",
                "llm": "LLMs",
                "phase_i_generation": "Generation PlantUML",
            }.items()
        }
        if pair.get("source_cells") != source_cells:
            raise RuntimeError(f"{pair_id} source_cells mismatch")
        expected_locator_data = {
            "sheet": SHEET,
            "data_row_index": data_row_index,
            "excel_row": excel_row,
            "selected_cell": selected["output"]["cell"],
        }
        if pair.get("source_locator_data") != expected_locator_data:
            raise RuntimeError(f"{pair_id} source_locator_data mismatch")
        if pair.get("source_locator_type") != "xlsx_feedback_stage_row_v2":
            raise RuntimeError(f"{pair_id} source_locator_type mismatch")
        expected_available = [
            {
                "column": stage["output"]["header"],
                "cell": stage["output"]["cell"],
                "sha256": stage["output"]["sha256"],
                "text_length": stage["output"]["char_length"],
            }
            for stage in lineage
            if stage["output"]["present"]
        ]
        if pair.get("available_stage_outputs") != expected_available:
            raise RuntimeError(f"{pair_id} available_stage_outputs mismatch")
        reference = source_row[_column_index(positions, "PlantUML")]
        nl_text = source_row[_column_index(positions, "Requirement Description")]
        if pair.get("reference_plantuml_sha256") != _sha256_text(reference):
            raise RuntimeError(f"{pair_id} reference hash mismatch")
        if pair.get("nl_sha256") != _sha256_text(nl_text):
            raise RuntimeError(f"{pair_id} NL hash mismatch")
    return {
        "trace_verified_pair_count": len(extracted),
        "source_workbook_sha256": workbook_sha256,
        "workbook_schema": schema,
        "workbook_schema_sha256": schema_sha256,
    }


def extract_feedback_final_pairs(
    workbook_path: Path,
    *,
    extracted_at: str,
    expected_pair_count: int = EXPECTED_PAIR_COUNT,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook_path = workbook_path.resolve()
    workbook_sha256 = _sha256_bytes(workbook_path.read_bytes())
    headers, positions, rows = _read_workbook_rows(workbook_path)
    if len(rows) != expected_pair_count:
        raise RuntimeError(
            f"expected {expected_pair_count} STM result rows, got {len(rows)}"
        )
    schema = _header_schema(headers)
    schema_sha256 = _sha256_text(_canonical_json(schema))

    extracted: list[dict[str, Any]] = []
    selected_stage_counts: Counter[str] = Counter()
    for index, (excel_row, source_row) in enumerate(rows):
        identity_values = {
            column: source_row[_column_index(positions, column)]
            for column in IDENTITY_COLUMNS
        }
        generation_index = _column_index(positions, "Generation PlantUML")
        generation_value = source_row[generation_index]
        missing_identity = [
            column
            for column, value in {
                **identity_values,
                "Generation PlantUML": generation_value,
            }.items()
            if _is_blank(value)
        ]
        if missing_identity:
            raise RuntimeError(
                f"STM Results excel row {excel_row} has empty required values: "
                f"{missing_identity}"
            )
        if any(
            not isinstance(value, str)
            for value in (*identity_values.values(), generation_value)
        ):
            raise RuntimeError(
                f"STM Results excel row {excel_row} has non-text identity values"
            )

        lineage = _stage_lineage(
            source_row,
            headers=headers,
            positions=positions,
            excel_row=excel_row,
        )
        selected = next(
            stage for stage in reversed(lineage) if stage["output"]["present"]
        )
        selected_column = selected["output"]["header"]
        selected_text = source_row[_column_index(positions, selected_column)]
        phase_i_text = generation_value
        pair_id = f"llms_emp_feedback_final_{index:04d}"
        selected_stage_counts[selected_column] += 1
        source_cells = {
            key: f"{get_column_letter(_column_index(positions, header) + 1)}{excel_row}"
            for key, header in {
                "model_source": "Model Source",
                "model_name": "Model Name",
                "nl": "Requirement Description",
                "reference_plantuml": "PlantUML",
                "llm": "LLMs",
                "phase_i_generation": "Generation PlantUML",
            }.items()
        }
        available_stages = [
            {
                "column": stage["output"]["header"],
                "cell": stage["output"]["cell"],
                "sha256": stage["output"]["sha256"],
                "text_length": stage["output"]["char_length"],
            }
            for stage in lineage
            if stage["output"]["present"]
        ]
        extracted.append(
            {
                "schema_version": PAIR_SCHEMA_VERSION,
                "pair_id": pair_id,
                "pair_set_id": "llms_emp_stm_feedback_final",
                "eligibility_state": "conditional_final_pool",
                "exclusion_reason": None,
                "extracted_at": extracted_at,
                "extraction_method": (
                    "openpyxl STM Results row extraction; select last sequential "
                    "author checking output with Semantic > Grammar > Format > Generation fallback"
                ),
                "selection_policy_id": SELECTION_POLICY_ID,
                "generation_actor": "LLM",
                "generation_context": "author_phase_ii_checking_feedback",
                "generation_model_or_method": (
                    f"{identity_values['LLMs']}; Experiment Results.xlsx / "
                    f"STM Results / {selected_column}"
                ),
                "is_generated_stm0": True,
                "is_postprocessed": selected["stage_id"] != "phase_i_generation",
                "is_reference": False,
                "llm": identity_values["LLMs"],
                "model_name": identity_values["Model Name"],
                "model_source": identity_values["Model Source"],
                "nl_role": "requirements_description",
                "nl_sha256": _sha256_text(identity_values["Requirement Description"]),
                "nl_source_column": "Requirement Description",
                "nl_text": identity_values["Requirement Description"],
                "phase_i_stm0_sha256": _sha256_text(phase_i_text),
                "phase_i_changed": selected_text != phase_i_text,
                "is_phase_i_fallback": selected["stage_id"] == "phase_i_generation",
                "reference_plantuml_sha256": _sha256_text(
                    identity_values["PlantUML"]
                ),
                "selected_stage": selected["stage_id"],
                "selected_stage_column": selected_column,
                "selected_stage_cell": selected["output"]["cell"],
                "stage_lineage": lineage,
                "available_stage_outputs": available_stages,
                "source_asset_id": "llms_emp_experiment_results_xlsx",
                "source_local_path": "raw/drive_download/Experiment Results.xlsx",
                "source_locator": (
                    f"sheet={SHEET}; row={index}; excel_row={excel_row}; "
                    f"selected_cell={selected['output']['cell']}; "
                    f"selected={selected_column}"
                ),
                "source_locator_type": "xlsx_feedback_stage_row_v2",
                "source_locator_data": {
                    "sheet": SHEET,
                    "data_row_index": index,
                    "excel_row": excel_row,
                    "selected_cell": selected["output"]["cell"],
                },
                "source_cells": source_cells,
                "source_row_index": index,
                "source_excel_row": excel_row,
                "source_sha256": workbook_sha256,
                "workbook_schema_sha256": schema_sha256,
                "stm0_role": "author_feedback_final_plantuml",
                "stm0_sha256": _sha256_text(selected_text),
                "stm0_source_column": selected_column,
                "stm0_text": selected_text,
                "stm_format": "plantuml",
                "trace_verified": False,
            }
        )

    verification = verify_feedback_final_trace(workbook_path, extracted)
    for row in extracted:
        row["trace_verified"] = True
    fallback_pair_ids = [
        row["pair_id"] for row in extracted if row["is_phase_i_fallback"]
    ]
    summary = {
        "schema_version": SUMMARY_SCHEMA_VERSION,
        "pair_schema_version": PAIR_SCHEMA_VERSION,
        "generated_at": extracted_at,
        "source_workbook": _repo_relative(workbook_path),
        "source_workbook_sha256": workbook_sha256,
        "source_sheet": SHEET,
        "workbook_schema": schema,
        "workbook_schema_sha256": schema_sha256,
        "selection_policy_id": SELECTION_POLICY_ID,
        "selection_priority": list(FINAL_STAGE_COLUMNS),
        "pair_count": len(extracted),
        "trace_verified_pair_count": verification["trace_verified_pair_count"],
        "selected_stage_counts": {
            column: selected_stage_counts.get(column, 0)
            for column in FINAL_STAGE_COLUMNS
        },
        "fallback_pair_ids": fallback_pair_ids,
        "phase_i_changed_count": sum(row["phase_i_changed"] for row in extracted),
        "phase_i_unchanged_count": sum(
            not row["phase_i_changed"] for row in extracted
        ),
        "unique_final_plantuml_count": len(
            {row["stm0_text"] for row in extracted}
        ),
        "llm_counts": dict(Counter(row["llm"] for row in extracted)),
    }
    return extracted, summary


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--summary", type=Path, default=DEFAULT_SUMMARY)
    parser.add_argument(
        "--extracted-at",
        default=datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
    )
    args = parser.parse_args()
    rows, summary = extract_feedback_final_pairs(
        args.workbook,
        extracted_at=args.extracted_at,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        "".join(
            json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in rows
        ),
        encoding="utf-8",
    )
    args.summary.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
