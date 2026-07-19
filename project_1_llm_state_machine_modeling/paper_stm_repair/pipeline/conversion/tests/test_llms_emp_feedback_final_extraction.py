from __future__ import annotations

import copy
import hashlib
import json
import sys
from collections import Counter
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def _repo_root() -> Path:
    for parent in Path(__file__).resolve().parents:
        if (parent / "project_1_llm_state_machine_modeling").is_dir():
            return parent
    raise RuntimeError("repository root not found")


REPO_ROOT = _repo_root()
PAPER_ROOT = REPO_ROOT / "project_1_llm_state_machine_modeling/paper_stm_repair"
TOOLS = PAPER_ROOT / "pipeline/conversion/tools"
sys.path.insert(0, str(TOOLS))

from extract_llms_emp_feedback_final import (  # noqa: E402
    FINAL_STAGE_COLUMNS,
    PAIR_SCHEMA_VERSION,
    SELECTION_POLICY_ID,
    SUMMARY_SCHEMA_VERSION,
    extract_feedback_final_pairs,
    verify_feedback_final_trace,
)


WORKBOOK = (
    PAPER_ROOT
    / "corpora/seed_library/llms-emp-stm-subset/assets/raw/drive_download"
    / "Experiment Results.xlsx"
)
EXTRACTED = (
    PAPER_ROOT
    / "corpora/seed_library/llms-emp-stm-subset/assets/extracted"
    / "feedback_final_pairs.jsonl"
)
SUMMARY = EXTRACTED.with_name("feedback_final_validation_summary.json")
FIXED_TIME = "2026-07-19T06:00:00+00:00"


def _sha256(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _plantuml(name: str) -> str:
    return f"@startuml\nstate {name}\n@enduml"


def _read_jsonl(path: Path) -> list[dict[str, object]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines()]


def _jsonl(rows: list[dict[str, object]]) -> str:
    return "".join(
        json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in rows
    )


def test_committed_feedback_final_pool_matches_every_workbook_row() -> None:
    extracted = _read_jsonl(EXTRACTED)
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=False)
    sheet = workbook["STM Results"]
    source_rows = sheet.iter_rows(values_only=True)
    headers = list(next(source_rows))
    column_index = {str(value): index for index, value in enumerate(headers) if value}
    rows = [
        row
        for row in source_rows
        if row[column_index["Requirement Description"]] is not None
        and str(row[column_index["Requirement Description"]]).strip()
    ]

    assert len(extracted) == len(rows) == 60
    assert [row["pair_id"] for row in extracted] == [
        f"llms_emp_feedback_final_{index:04d}" for index in range(60)
    ]
    assert Counter(row["llm"] for row in extracted) == {
        "Claude": 10,
        "DeepSeek": 10,
        "GPT-4": 10,
        "GPT-4o": 10,
        "Kimi": 10,
        "Llama": 10,
    }

    for index, (source_row, extracted_row) in enumerate(zip(rows, extracted)):
        values = {
            column: source_row[column_index[column]]
            for column in FINAL_STAGE_COLUMNS
        }
        selected = next(
            column
            for column in FINAL_STAGE_COLUMNS
            if values[column] is not None and str(values[column]).strip()
        )
        selected_text = str(values[selected])
        lineage = extracted_row["stage_lineage"]

        assert extracted_row["schema_version"] == PAIR_SCHEMA_VERSION
        assert extracted_row["selection_policy_id"] == SELECTION_POLICY_ID
        assert extracted_row["source_row_index"] == index
        assert extracted_row["source_excel_row"] == index + 2
        assert extracted_row["selected_stage_column"] == selected
        assert extracted_row["selected_stage_cell"].endswith(str(index + 2))
        assert extracted_row["stm0_text"] == selected_text
        assert extracted_row["stm0_sha256"] == _sha256(selected_text)
        assert len(lineage) == 4
        assert [stage["output"]["column_letter"] for stage in lineage] == [
            "I",
            "U",
            "Z",
            "AE",
        ]
        assert all(stage["output"]["cell"].endswith(str(index + 2)) for stage in lineage)
        assert extracted_row["trace_verified"] is True
        assert extracted_row["is_reference"] is False
    workbook.close()

    assert [row["pair_id"] for row in extracted if row["is_phase_i_fallback"]] == [
        "llms_emp_feedback_final_0054",
        "llms_emp_feedback_final_0055",
    ]
    for pair_id in ("llms_emp_feedback_final_0054", "llms_emp_feedback_final_0055"):
        row = next(item for item in extracted if item["pair_id"] == pair_id)
        assert [stage["output"]["present"] for stage in row["stage_lineage"]] == [
            True,
            False,
            False,
            False,
        ]
        assert all(
            stage["output"]["cell"] is not None for stage in row["stage_lineage"]
        )


def test_committed_pool_has_independently_verified_trace() -> None:
    extracted = _read_jsonl(EXTRACTED)
    result = verify_feedback_final_trace(WORKBOOK, extracted)

    assert result["trace_verified_pair_count"] == 60
    assert result["source_workbook_sha256"] == extracted[0]["source_sha256"]
    assert result["workbook_schema_sha256"] == extracted[0][
        "workbook_schema_sha256"
    ]


def test_committed_feedback_final_pool_is_deterministically_reproducible() -> None:
    rows_a, summary_a = extract_feedback_final_pairs(WORKBOOK, extracted_at=FIXED_TIME)
    rows_b, summary_b = extract_feedback_final_pairs(WORKBOOK, extracted_at=FIXED_TIME)

    assert rows_a == rows_b == _read_jsonl(EXTRACTED)
    assert _jsonl(rows_a) == _jsonl(rows_b) == EXTRACTED.read_text(encoding="utf-8")
    assert summary_a == summary_b == json.loads(SUMMARY.read_text(encoding="utf-8"))
    assert summary_a["schema_version"] == SUMMARY_SCHEMA_VERSION
    assert summary_a["pair_schema_version"] == PAIR_SCHEMA_VERSION
    assert summary_a["selected_stage_counts"] == {
        "Generation PlantUML": 2,
        "Result with Format Checking": 0,
        "Result with Grammar Checking": 0,
        "Result with Semantic Checking": 58,
    }
    assert summary_a["fallback_pair_ids"] == [
        "llms_emp_feedback_final_0054",
        "llms_emp_feedback_final_0055",
    ]
    assert summary_a["phase_i_changed_count"] == 52
    assert summary_a["phase_i_unchanged_count"] == 8
    assert summary_a["trace_verified_pair_count"] == 60


def _write_fixture(
    path: Path,
    *,
    blank_generation_row: int | None = None,
    invalid_stage: tuple[int, str] | None = None,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "STM Results"
    headers = [
        "Model Source",
        "Model Name",
        "Requirement Description",
        "PlantUML",
        "LLMs",
        "Generation PlantUML",
        "Format Hallucinations",
        "Result with Format Checking",
        "Resolved",
        "SysML Grammar Hallucinations",
        "Result with Grammar Checking",
        "Resolved",
        "Semmantic Hallucinations",
        "Result with Semantic Checking",
        "Visualisation",
        "Resolved",
    ]
    sheet.append(headers)
    stage_rows = [
        (_plantuml("generation0"), None, None, None),
        (_plantuml("generation1"), _plantuml("format1"), None, None),
        (
            _plantuml("generation2"),
            _plantuml("format2"),
            _plantuml("grammar2"),
            None,
        ),
        (
            _plantuml("generation3"),
            _plantuml("format3"),
            _plantuml("grammar3"),
            _plantuml("semantic3"),
        ),
    ]
    stage_columns = {
        "Generation PlantUML": 0,
        "Result with Format Checking": 1,
        "Result with Grammar Checking": 2,
        "Result with Semantic Checking": 3,
    }
    for index, source_stages in enumerate(stage_rows):
        stages = list(source_stages)
        if blank_generation_row == index:
            stages[0] = None
        if invalid_stage is not None and invalid_stage[0] == index:
            stages[stage_columns[invalid_stage[1]]] = "not PlantUML"
        generation, fmt, grammar, semantic = stages
        sheet.append(
            [
                "source",
                f"model-{index}",
                f"requirement-{index}",
                _plantuml(f"reference{index}"),
                "LLM",
                generation,
                "format issue" if fmt else None,
                fmt,
                "YES" if fmt else None,
                "grammar issue" if grammar else None,
                grammar,
                "YES" if grammar else None,
                "None" if semantic else None,
                semantic,
                None,
                1 if semantic else None,
            ]
        )
    workbook.save(path)


def test_stage_priority_uses_the_last_sequential_regeneration(tmp_path: Path) -> None:
    workbook = tmp_path / "stages.xlsx"
    _write_fixture(workbook)
    rows, summary = extract_feedback_final_pairs(
        workbook,
        extracted_at=FIXED_TIME,
        expected_pair_count=4,
    )

    assert [row["selected_stage"] for row in rows] == [
        "phase_i_generation",
        "phase_ii_format",
        "phase_ii_grammar",
        "phase_ii_semantic",
    ]
    assert [row["selected_stage_cell"] for row in rows] == ["F2", "H3", "K4", "N5"]
    assert [row["stm0_text"] for row in rows] == [
        _plantuml("generation0"),
        _plantuml("format1"),
        _plantuml("grammar2"),
        _plantuml("semantic3"),
    ]
    assert [stage["output"]["present"] for stage in rows[0]["stage_lineage"]] == [
        True,
        False,
        False,
        False,
    ]
    assert rows[3]["stage_lineage"][3]["feedback"]["value"] == "None"
    assert rows[3]["stage_lineage"][3]["resolved"]["value"] == 1
    assert summary["pair_count"] == 4


def test_missing_phase_i_generation_fails_closed(tmp_path: Path) -> None:
    workbook = tmp_path / "missing-generation.xlsx"
    _write_fixture(workbook, blank_generation_row=0)
    with pytest.raises(RuntimeError, match="Generation PlantUML"):
        extract_feedback_final_pairs(
            workbook,
            extracted_at=FIXED_TIME,
            expected_pair_count=4,
        )


@pytest.mark.parametrize(
    "column",
    [
        "Generation PlantUML",
        "Result with Format Checking",
        "Result with Grammar Checking",
        "Result with Semantic Checking",
    ],
)
def test_any_nonempty_non_plantuml_stage_fails_closed(
    tmp_path: Path, column: str
) -> None:
    workbook = tmp_path / f"invalid-{column.replace(' ', '-')}.xlsx"
    _write_fixture(workbook, invalid_stage=(3, column))

    with pytest.raises(RuntimeError, match="non-PlantUML stage output"):
        extract_feedback_final_pairs(
            workbook,
            extracted_at=FIXED_TIME,
            expected_pair_count=4,
        )


@pytest.mark.parametrize(
    ("field", "bad_value"),
    [
        ("stm0_text", _plantuml("mutated")),
        ("stm0_sha256", "0" * 64),
        ("selected_stage", "phase_ii_semantic"),
        ("selected_stage_cell", "AE999"),
        ("is_phase_i_fallback", False),
        ("source_excel_row", 999),
        ("workbook_schema_sha256", "0" * 64),
    ],
)
def test_independent_trace_verifier_rejects_pair_mutations(
    tmp_path: Path, field: str, bad_value: object
) -> None:
    workbook = tmp_path / "trace.xlsx"
    _write_fixture(workbook)
    rows, _ = extract_feedback_final_pairs(
        workbook,
        extracted_at=FIXED_TIME,
        expected_pair_count=4,
    )
    mutated = copy.deepcopy(rows)
    mutated[0][field] = bad_value

    with pytest.raises(RuntimeError, match="mismatch"):
        verify_feedback_final_trace(workbook, mutated)


def test_independent_trace_verifier_rejects_lineage_mutation(tmp_path: Path) -> None:
    workbook = tmp_path / "lineage.xlsx"
    _write_fixture(workbook)
    rows, _ = extract_feedback_final_pairs(
        workbook,
        extracted_at=FIXED_TIME,
        expected_pair_count=4,
    )
    mutated = copy.deepcopy(rows)
    mutated[1]["stage_lineage"][1]["output"]["sha256"] = "0" * 64

    with pytest.raises(RuntimeError, match="stage lineage mismatch"):
        verify_feedback_final_trace(workbook, mutated)


def test_independent_trace_verifier_rejects_header_schema_drift(tmp_path: Path) -> None:
    workbook = tmp_path / "schema.xlsx"
    _write_fixture(workbook)
    rows, _ = extract_feedback_final_pairs(
        workbook,
        extracted_at=FIXED_TIME,
        expected_pair_count=4,
    )

    changed = load_workbook(workbook)
    changed["STM Results"]["O1"] = "Changed Visualisation Header"
    changed.save(workbook)
    changed.close()
    mutated = copy.deepcopy(rows)
    workbook_sha256 = hashlib.sha256(workbook.read_bytes()).hexdigest()
    for row in mutated:
        row["source_sha256"] = workbook_sha256

    with pytest.raises(RuntimeError, match="schema fingerprint mismatch"):
        verify_feedback_final_trace(workbook, mutated)
